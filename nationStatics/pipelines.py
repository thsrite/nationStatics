# -*- coding: utf-8 -*-

import pymysql

from openpyxl import Workbook


class NationstaticsPipeline(object):

    conn = None
    cue = None

    def __init__(self):
        # 租房主表
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.append(
            ['id', 'title', 'gid', 'catid', 'catname', 'areaid', 'begintime', 'activetime', 'endtime', 'content',
             'userid',
             'contact_who', 'qq', 'email', 'tel', 'updatetime', 'hit', 'ismember', 'manage_pwd', 'ip', 'ip2area',
             'info_level', 'img_path', 'img_count', 'upgrade_type', 'upgrade_time', 'upgrade_type_list',
             'upgrade_time_list', 'ifred', 'ifbold', 'certify', 'dir_typename', 'html_path', 'upgrade_type_index',
             'upgrade_time_index', 'mappoint', 'web_address', 'latitude', 'longitude'])
        # 租房副表
        self.wb1 = Workbook()
        self.ws1 = self.wb1.active
        self.ws1.append(['iid', 'id', 'position', 'rent_type', 'room_type', 'mini_rent', 'content', 'house_pro'])
        # 租房副副表
        self.wb2 = Workbook()
        self.ws2 = self.wb2.active
        self.ws2.append(['image_id', 'path', 'prepath', 'infoid', 'uptime'])

        #手机副表
        self.wb3 = Workbook()
        self.ws3 = self.wb3.active
        self.ws3.append(['id','mbrand','price','new_old','from'])

        #招聘副表
        self.wb4 = Workbook()
        self.ws4 = self.wb4.active
        self.ws4.append(['id','sex_demand','salary','job','company','content','education','work_life','fuli','property'])

        #二手车副表
        self.wb5 = Workbook()
        self.ws5 = self.wb5.active
        self.ws5.append(['id', 'car_brand', 'car_year', 'mileage', 'prices', 'content', 'new_old'])

        # 数据库连接
        self.con = pymysql.connect(host='47.105.207.207', user='root', passwd='root', db='app', charset='utf8', port=3306)
        # 数据库游标

        self.cue = self.con.cursor()

    def process_item(self, item, spider):
        if item['type'] == "房屋出租":
            # line = [item['begintime'], item['title'], '3', '41', '房屋出租', '1', item['begintime'], '0', '0', item['desc'],
            #         '0', item['man'], '',
            #         '', item['phone'], '0', '0', '0', 'c4ca4238a0b923820dcc509a6f75849b', '223.81.192.35', '山东省青岛市移动', '1',
            #         item['preUrlArr'][0], item['num'], '1', '0', '1', '0', '0', '0', '0', 'chuzu', '', '1', '0', '', '', '',
            #         '']
            # self.ws.append(line)
            # self.wb.save('information.xlsx')
            #
            # line1 = [item['begintime'], item['begintime'], '1', '1', item["room_type"], item["price"], '', '1,2,3,4,5']
            # self.ws1.append(line1)
            # self.wb1.save('information23.xlsx')
            #
            # for i in range(0, len(item['urlArr'])):
            #     line2 = [item['numArr'][i], item['urlArr'][i], item['preUrlArr'][i], item['begintime'], item['begintime']]
            #     self.ws2.append(line2)
            # self.wb2.save('my_info_img.xlsx')

            #dict(item)
            # 数据库连接
            self.con = pymysql.connect(host='*.*.*.*', user='root', passwd='root', db='app', charset='utf8', port=3306)
            print("mysql connect succes")  # 测试语句，这在程序执行时非常有效的理解程序是否执行到这一步

            try:
                self.cue.execute("INSERT INTO my_information() VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
                ,[item['begintime'], item['title'], '3', '41', '房屋出租', '1', item['begintime'], '0', '0',
                        item['desc'], '0', item['man'], '',
                        '', item['phone'], '0', '0', '0', 'c4ca4238a0b923820dcc509a6f75849b', '223.81.192.35',
                        '山东省青岛市移动', '1',
                        item['preUrlArr'][0], item['num'], '1', '0', '1', '0', '0', '0', '0', 'chuzu', '', '1', '0', '', '',
                        0,
                        0])
                print("插入主表成功")

                self.cue.execute("INSERT INTO my_information_23() VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                [item['begintime'], item['begintime'], '1', '1', item["room_type"], item["price"], '', '1,2,3,4,5'])
                print("插入副表成功")
                for i in range(0, len(item['urlArr'])):
                    self.cue.execute("insert into my_info_img(image_id,path,prepath,infoid,uptime) values (%s,%s,%s,%s,%s)",[item['numArr'][i], item['urlArr'][i], item['preUrlArr'][i], item['begintime'], item['begintime']])
                print("插入副副表成功")

                print("insert success")  # 测试语句
            except Exception as e:
                print('Insert error:', e)
                self.con.rollback()
            else:
                self.con.commit()

        if item["type"] == "二手手机":
            # line = [item['begintime'], item['title'], '1', '11', '手机转让', '1', item['begintime'], '0', '0', item['desc'],
            #         '0', item['man'], '',
            #         '', item['phone'], '0', '0', '0', 'c4ca4238a0b923820dcc509a6f75849b', '223.81.192.35', '山东省青岛市移动',
            #         '1',
            #         item['preUrlArr'][0], item['num'], '1', '0', '1', '0', '0', '0', '0', 'shouji', '', '1', '0', '', '',
            #         '',
            #         '']
            # self.ws.append(line)
            # self.wb.save('information.xlsx')
            #
            # line1 = [item['begintime'],item['pinpai'],item['price'],item['new_old'],'1']
            # self.ws3.append(line1)
            # self.wb3.save('information28.xlsx')
            #
            # for i in range(0, len(item['urlArr'])):
            #     line2 = [item['numArr'][i], item['urlArr'][i], item['preUrlArr'][i], item['begintime'],
            #              item['begintime']]
            #     self.ws2.append(line2)
            # self.wb2.save('my_info_img.xlsx')

            # dict(item)
            # 数据库连接
            self.con = pymysql.connect(host='*.*.*.*', user='root', passwd='root', db='app', charset='utf8',
                                       port=3306)
            print("mysql connect succes")  # 测试语句，这在程序执行时非常有效的理解程序是否执行到这一步

            try:
                self.cue.execute(
                    "INSERT INTO my_information() VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
                    , [item['begintime'], item['title'], '1', '11', '手机转让', '1', item['begintime'], '0', '0',
                       item['desc'], '0', item['man'], '',
                       '', item['phone'], '0', '0', '0', 'c4ca4238a0b923820dcc509a6f75849b', '223.81.192.35',
                       '山东省青岛市移动', '1',
                       item['preUrlArr'][0], item['num'], '1', '0', '1', '0', '0', '0', '0', 'shouji', '', '1', '0', '',
                       '',
                       0,
                       0])
                print("插入主表成功")

                self.cue.execute("INSERT INTO my_information_28(id,mbrand,price,new_old) VALUES (%s,%s,%s,%s)",
                                 [item['begintime'],item['pinpai'],item['price'],item['new_old']])
                print("插入副表成功")
                for i in range(0, len(item['urlArr'])):
                    self.cue.execute(
                        "insert into my_info_img(image_id,path,prepath,infoid,uptime) values (%s,%s,%s,%s,%s)",
                        [item['numArr'][i], item['urlArr'][i], item['preUrlArr'][i], item['begintime'],
                         item['begintime']])
                print("插入副副表成功")

                print("insert success")  # 测试语句
            except Exception as e:
                print('Insert error:', e)
                self.con.rollback()
            else:
                self.con.commit()

        if item["type"] == "招聘":
            # line = [item['begintime'], item['title'], '4', item["zhiye"], item['zhiyename'], '1', item['begintime'], '0', '0', item['desc'],
            #         '0', item['contryname'] , '',
            #         '', item['phone'], '0', '0', '0', 'c4ca4238a0b923820dcc509a6f75849b', '223.81.192.35', '山东省青岛市移动',
            #         '1',
            #         '', 0, '1', '0', '1', '0', '0', '0', '0',item["type"], '', '1', '0', '', '',
            #         '',
            #         '']
            # self.ws.append(line)
            # self.wb.save('information.xlsx')
            #
            # line1 = [item['begintime'],'1,2',item['price'],item['title'],item['contryname'],'',item['xueli'],item['work_life'],item['fuli'],1]
            # self.ws4.append(line1)
            # self.wb4.save('information7.xlsx')

            # dict(item)
            # 数据库连接
            self.con = pymysql.connect(host='*.*.*.*', user='root', passwd='root', db='app', charset='utf8',
                                       port=3306)
            print("mysql connect succes")  # 测试语句，这在程序执行时非常有效的理解程序是否执行到这一步

            try:
                self.cue.execute(
                    "INSERT INTO my_information() VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
                    , [item['begintime'], item['title'], '4', item["zhiye"], item['zhiyename'], '1', item['begintime'], '0', '0', item['desc'],
                    '0', item['contryname'], '',
                    '', item['phone'], '0', '0', '0', 'c4ca4238a0b923820dcc509a6f75849b', '223.81.192.35', '山东省青岛市移动',
                    '1',
                    '', 0, '1', '0', '1', '0', '0', '0', '0',item["type"], '', '1', '0', '', '',
                    0,
                    0])
                print("插入主表成功")

                self.cue.execute("INSERT INTO my_information_7(id,sex_demand,salary,job,company,content,education,work_life,fuli,property) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                                 [item['begintime'],'1,2',item['price'],item['title'],item['contryname'],'',item['xueli'],item['work_life'],item['fuli'],1])
                print("插入副表成功")

                print("insert success")  # 测试语句
            except Exception as e:
                print('Insert error:', e)
                self.con.rollback()
            else:
                self.con.commit()

        if item["type"] == "二手车":
            # line = [item['begintime'], item['title'], '2', '28', '二手轿车', '1', item['begintime'], '0', '0', item['desc'],
            #     '0', item['man'], '',
            #     '', item['phone'], '0', '0', '0', 'c4ca4238a0b923820dcc509a6f75849b', '223.81.192.35', '山东省青岛市移动', '1',
            #     item['imgUrl'], item['num'], '1', '0', '1', '0', '0', '0', '0', 'jiaoche', '', '1', '0', '', '', '', '']
            # self.ws.append(line)
            # self.wb.save('information.xlsx')
            #
            # line1 = [item['begintime'], item['pinpai'],item['che_year'],item['gongli'],item["price"],'',item["new_old"]]
            # self.ws5.append(line1)
            # self.wb5.save('information12.xlsx')
            #
            # for i in range(0, len(item['urlArr'])):
            #     line2 = [item['numArr'][i], item['urlArr'][i], item['preUrlArr'][i], item['begintime'],
            #              item['begintime']]
            #     self.ws2.append(line2)
            # self.wb2.save('my_info_img.xlsx')

            # dict(item)
            # 数据库连接
            self.con = pymysql.connect(host='*.*.*.*', user='root', passwd='root', db='app', charset='utf8',
                                       port=3306)
            print("mysql connect succes")  # 测试语句，这在程序执行时非常有效的理解程序是否执行到这一步

            try:
                self.cue.execute(
                    "INSERT INTO my_information() VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
                    , [item['begintime'], item['title'], '2', '28', '二手轿车', '1', item['begintime'], '0', '0', item['desc'],
                '0', item['man'], '',
                '', item['phone'], '0', '0', '0', 'c4ca4238a0b923820dcc509a6f75849b', '223.81.192.35', '山东省青岛市移动', '1',
                item['preUrlArr'][0], item['num'], '1', '0', '1', '0', '0', '0', '0', 'jiaoche', '', '1', '0', '', '', 0, 0])
                print("插入主表成功")

                self.cue.execute("INSERT INTO my_information_12(id, car_brand, car_year, mileage, prices,content, new_old) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                                 [item['begintime'], item['pinpai'], item['che_year'], item['gongli'], item["price"],
                                  '', item["new_old"]])
                print("插入副表成功")
                for i in range(0, len(item['urlArr'])):
                    self.cue.execute(
                        "insert into my_info_img(image_id,path,prepath,infoid,uptime) values (%s,%s,%s,%s,%s)",
                        [item['numArr'][i], item['urlArr'][i], item['preUrlArr'][i], item['begintime'],
                         item['begintime']])
                print("插入副副表成功")

                print("insert success")  # 测试语句
            except Exception as e:
                print('Insert error:', e)
                self.con.rollback()
            else:
                self.con.commit()

        self.con.close()
        return item
